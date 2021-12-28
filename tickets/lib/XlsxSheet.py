from openpyxl import load_workbook
import xlsxwriter

setting = {
    # 用于转换一些中文的字段名转换成英文，数据在输出前会修改字典返回的键名,默认转换是从第一个字段开始
    'CHANGE_FIELD': ('sys_sku','mum_id',)
}
# 操作xlsx文档某一页的类
class XlsxSheet:
    # 初始化
    # @ workBookPath str 工作薄的地址，绝对位置或者相对位置
    # @ sheet  第几个工作薄，以数字从0开始，
    def __init__(self, load_workbook, workBookPath, sheetNum):
        self.workBookPath = workBookPath
        self.sheetNum = sheetNum
        self.wb = load_workbook(workBookPath)
        self.sheets = self.wb.worksheets
        self.sheet = self.sheets[sheetNum]
        # 定义一个字典用来存放一行数据
        self.data = {}
        self.field = self.get_field()
        # 设置文件
        #


    # 根据行数查询一行记录
    def get_row_by_id(self, rowNum = 2):
        data = self.sheet[rowNum]
        lis = []
        for row in data:
            lis.append(row.value)

        self.data = self.format_data(self.field,lis)
        return self.data

    # 取得行数据中的cell 对象的value值返回一个list
    # @param data Tuple 元祖对象
    def get_cell_values(self,data):
        lis = []
        for row in data:
            lis.append(row.value)
        return lis

    # 根据条件查询多行
    # @param dic dict 包含条件的字典数据
    # return list or null 符合参数dic 键名 = 值的数据会被以list的形式返回
    def get_many_rows(self, dic):
        datas = []
        for row in self.sheet:
            d = self.get_cell_values(row)
            data = self.format_data(self.field, d)
            boo = 0
            for i in dic.items():
                if(i[1] != data[i[0]]):
                    boo = 1
            if boo != 1:
                # print(data)
                datas.append(data)
        return datas

    # 添加一行到文本尾部
    # @param list 添加的一行从A1开始按照list顺序添加
    def append_row(self,data):
        # 需要打开文档的可编辑权限，要不调用该函数就会报错
        # 建文件及sheet.
        # 获取当前活跃的sheet，默认是第一个sheet
        ws = self.wb.active
        # ws['A1'] = 'class'
        # ws['B1'] = 'name'
        # ws['C1'].value = 'score'
        # row1 = ['class1', 'zhangsan', 90]
        # row2 = ['class2', 'lisi', 88]
        # ws.append(row1)
        ws.append(data)
        self.wb.save(self.workBookPath)

    # 取字段数据，默认取当前sheet的第一行数据，一般为字段名
    # @ num 取第几行元素默认为1
    # @return list 返回一个包含第一行内容的list
    def get_field(self, num = 1):
        li = []
        data = self.sheet[num]
        for row in data:
            li.append(row.value)

        # 转换字段名
        for i in range(len(setting['CHANGE_FIELD'])):
            li[i] = setting['CHANGE_FIELD'][i]

        return li

    # 格式化数据,合并两个list 转换成一个字典
    # @param keys list 作为键名的list
    # @param data list 作为值的list
    # @return dict
    def format_data(self, keys, data):
        dic = {}
        for i in range(len(keys)):
            dic[keys[i]] = data[i]
        return dic

    # 更新数据
    # @ param
    def update_row(self):
        pass

    # 删除一行数据
    def delete_row(self):
        pass

def main():
    xlsxSheet = XlsxSheet(load_workbook, '图片资料和供应商(1).xlsx', 0)
    # 取表头字段
    # print(xlsxSheet.get_field())

    #取数据
    # print(xlsxSheet.get_row_by_id(4))

    # 查询符合条件的数据
    lis = xlsxSheet.get_many_rows({'sys_sku':'3179347'})
    print(lis)
    xlsxSheet.append_row(['1222111','4556554456'])
    # sublime 不支持repr 输出
    # repr(lis)
if __name__ == '__main__':
    main()