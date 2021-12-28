from openpyxl import load_workbook
from lib.XlsxSheet import XlsxSheet
# openpyxl 是支持xlsx 读写的工具。性能优于   xlrd&xlwt，xlsxwrite 只支持写操作
wb = load_workbook('guotaijunan.xlsx')
sheets = wb.worksheets   # 获取当前所有的sheet
print(sheets)

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

rows = sheet1.rows
columns = sheet1.columns
print(rows)
print(columns)
print('')

xlsxSheet = XlsxSheet(load_workbook, 'guotaijunan.xlsx', 0)
xlsxSheet.
# 迭代读取所有的行
for row in rows:
    print(row)
    row_val = [col.value for col in row]
    print(row_val)
print('')

# 迭代读取所有的列
for col in columns:
    print("col:",col)
    col_val = [row.value for row in col]
    print(col_val)